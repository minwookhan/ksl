from openpyxl import load_workbook
from tqdm import tqdm  # 진행률 표시

def delete_rows_based_on_conditions(file_path, sheet_name):
    """
    특정 엑셀 파일에서 5번째 컬럼(E)이 1이 아니면 삭제, 28번째 컬럼(AB)이 0이면 삭제 (진행률 표시 추가).

    :param file_path: 엑셀 파일 경로
    :param sheet_name: 처리할 시트 이름
    """
    print("📂 엑셀 파일을 로드하는 중...")
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    print(f"✅ '{file_path}' 파일의 '{sheet_name}' 시트에서 행 삭제를 시작합니다.\n")

    # 모든 데이터를 한 번에 가져와서 리스트로 변환
    data = list(ws.iter_rows(values_only=True))
    total_rows = len(data) - 1  # 헤더 제외

    print(f"📊 총 {total_rows}개의 행을 처리합니다...\n")

    # 진행률 표시 + 필터링 수행
    filtered_data = [data[0]]  # 헤더 유지
    for row in tqdm(data[1:], desc="Processing Rows", unit="row", total=total_rows):
        if row[4] == 1 and row[27] != 0:  # 조건: (E 컬럼 == 1) and (AB 컬럼 != 0)
            filtered_data.append(row)

    # 기존 데이터 삭제
    ws.delete_rows(2, ws.max_row)  # 헤더(1행) 제외하고 전체 삭제

    # 필터링된 데이터 다시 추가
    for row in tqdm(filtered_data[1:], desc="Writing Data", unit="row", total=len(filtered_data) - 1):
        ws.append(row)

    wb.save(file_path)
    wb.close()
   
    print(f"📊 최종 데이터 행 개수: {len(filtered_data) - 1}개 (헤더 제외)\n")


